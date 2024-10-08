from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import random
import math


# (1)
critiques = {
    'Lisa Rose': {'Lady': 2.5, 'Snakes': 3.5, 'Luck': 3.0, 'Superman': 3.5, 'Dupree': 2.5, 'Night': 3.0},
    'Gene Seymour': {'Lady': 3.0, 'Snakes': 3.5, 'Luck': 1.5, 'Superman': 5.0, 'Dupree': 3.5, 'Night': 3.0},
    'Michael Phillips': {'Lady': 2.5, 'Snakes': 3.0, 'Superman': 3.5, 'Night': 4.0},
    'Claudia Puig': {'Snakes': 3.5, 'Luck': 3.0, 'Superman': 4.0, 'Dupree': 2.5, 'Night': 4.5},
    'Mick Lasalle': {'Lady': 3.0, 'Snakes': 4.0, 'Luck': 2.0, 'Superman': 3.0, 'Dupree': 2.0, 'Night': 3.0},
    'Jack Matthews': {'Lady': 3.0, 'Snakes': 4.0, 'Superman': 5.0, 'Dupree': 3.5, 'Night': 3.0},
    'Toby': {'Snakes': 4.5, 'Superman': 4.0, 'Dupree': 1.0},
    'Anne': {'Lady': 1.5, 'Luck': 4.0, 'Dupree': 2.0}
}

# 2(a)(i)
def sim_distanceManhattan(person1, person2):
    """Calcule la distance de Manhattan en ignorant les valeurs 'None'."""
    communs = {item for item in person1 if item in person2 and person1[item] is not None and person2[item] is not None}

    if len(communs) == 0:
        return 0

    distance = sum(abs(person1[item] - person2[item]) for item in communs)
    return distance


def sim_distanceEuclidienne(person1, person2):
    """Calcule la distance Euclidienne en ignorant les valeurs 'None'."""
    communs = {item for item in person1 if item in person2 and person1[item] is not None and person2[item] is not None}

    if len(communs) == 0:
        return 0

    sum_of_squares = sum((person1[item] - person2[item]) ** 2 for item in communs)
    return math.sqrt(sum_of_squares)


# Test des fonctions de distance
print("Distance Manhattan entre Lisa Rose et Gene Seymour:",
      sim_distanceManhattan(critiques['Lisa Rose'], critiques['Gene Seymour']))
print("Distance Euclidienne entre Lisa Rose et Gene Seymour:",
      sim_distanceEuclidienne(critiques['Lisa Rose'], critiques['Gene Seymour']))


# 2(a)(ii)
def computeNearestNeighbor(nouveauCritique, critiques):
    """Retourne une liste triée des critiques proches de nouveauCritique basée sur la distance de Manhattan."""
    distances = []
    for critique in critiques:
        if critique != nouveauCritique:
            distance = sim_distanceManhattan(critiques[critique], critiques[nouveauCritique])
            distances.append((distance, critique))

    # Tri des distances par ordre croissant (les critiques les plus proches en premier)
    distances.sort()
    return distances


def recommend(nouveauCritique, critiques):
    """Recommande des films à l'utilisateur nouveauCritique en fonction des goûts des autres critiques."""
    # Trouver le critique le plus proche de nouveauCritique
    nearest = computeNearestNeighbor(nouveauCritique, critiques)[0][1]

    # Créer une liste pour stocker les recommandations
    recommendations = []

    # Parcourir les films vus par le critique le plus proche
    for film in critiques[nearest]:
        # Si le nouveau critique n'a pas vu le film et que le critique proche a donné une note
        if film not in critiques[nouveauCritique] and critiques[nearest][film] > 0:
            # Ajouter le film et la note donnée par le critique proche à la liste des recommandations
            recommendations.append((film, critiques[nearest][film]))

    # Trier les recommandations par la note, les films les mieux notés en premier
    recommendations.sort(key=lambda x: x[1])

    return recommendations


# Test de la recommandation pour Toby
recommandations = recommend('Toby', critiques)
print("Film recommandé à Toby :", recommandations)


# 2(b)(i)
def Bestrecommend(nouveauCritique, critiques):
    """Propose une recommandation à Anne en utilisant un score global pondéré."""
    totals = {}
    sim_sums = {}

    # Parcourir tous les critiques
    for critique in critiques:
        # Ignorer Anne elle-même
        if critique == nouveauCritique:
            continue

        # Calculer la distance de Manhattan entre Anne et le critique
        distance = sim_distanceManhattan(critiques[critique], critiques[nouveauCritique])

        # Ignorer si la distance est de 0 (pas de films en commun)
        if distance == 0:
            continue

        # Pondération de la contribution de ce critique à Anne
        weight = 1 / (1 + distance)

        # Parcourir les films notés par ce critique
        for film in critiques[critique]:
            # Seuls les films non vus par Anne sont intéressants
            if (film not in critiques[nouveauCritique] or critiques[nouveauCritique][film] == 0 or
                critiques[nouveauCritique][film] is None) and critiques[critique][film] is not None:
                # Ajouter le score pondéré au total
                if film not in totals:
                    totals[film] = 0
                totals[film] += critiques[critique][film] * weight

                # Ajouter au dénominateur
                if film not in sim_sums:
                    sim_sums[film] = 0
                sim_sums[film] += weight

    # Calculer la liste finale des recommandations
    rankings = [(total / sim_sums[film], film) for film, total in totals.items()]

    # Trier les recommandations par score décroissant
    rankings.sort(reverse=True)

    return rankings


# Test de la recommandation pondérée pour Anne
recommandations = Bestrecommend('Anne', critiques)
print("Film recommandé à Anne avec score pondéré :", recommandations[0])


# 2(b)(ii)
def OtherBestrecommend(nouveauCritique, critiques):
    """Propose une recommandation en utilisant un score global pondéré avec 1 / (1 + distance^2)."""
    totals = {}
    sim_sums = {}

    # Parcourir tous les critiques
    for critique in critiques:
        # Ignorer le critique lui-même (Anne)
        if critique == nouveauCritique:
            continue

        # Calculer la distance de Manhattan entre Anne et le critique
        distance = sim_distanceManhattan(critiques[critique], critiques[nouveauCritique])

        # Ignorer si la distance est de 0 (pas de films en commun)
        if distance == 0:
            continue

        # Pondération de la contribution avec la nouvelle formule
        weight = 1 / (1 + distance ** 2)

        # Parcourir les films notés par ce critique
        for film in critiques[critique]:
            # Seuls les films non vus par Anne sont intéressants
            if (film not in critiques[nouveauCritique] or critiques[nouveauCritique][film] == 0 or
                critiques[nouveauCritique][film] is None) and critiques[critique][film] is not None:
                # Ajouter le score pondéré au total
                if film not in totals:
                    totals[film] = 0
                totals[film] += critiques[critique][film] * weight

                # Ajouter au dénominateur
                if film not in sim_sums:
                    sim_sums[film] = 0
                sim_sums[film] += weight

    # Calculer la liste finale des recommandations
    rankings = [(total / sim_sums[film], film) for film, total in totals.items()]

    # Trier les recommandations par score décroissant
    rankings.sort(reverse=True)

    return rankings


# Test de la recommandation pondérée avec 1 / (1 + distance^2)
recommandations = OtherBestrecommend('Anne', critiques)
print("Film recommandé à Anne avec la nouvelle pondération :", recommandations[0])


# (3)
def pearson(person1, person2):
    communs = [film for film in person1 if film in person2 and person1[film] is not None and person2[film] is not None]

    n = len(communs)

    # If no common films, return 0
    if n == 0:
        return 0

    sum_xy = sum_x = sum_y = sum_x2 = sum_y2 = 0

    for film in communs:
        x = person1[film]
        y = person2[film]
        sum_xy += x * y
        sum_x += x
        sum_y += y
        sum_x2 += x ** 2
        sum_y2 += y ** 2

    denominator_x = sum_x2 - (sum_x ** 2) / n
    denominator_y = sum_y2 - (sum_y ** 2) / n

    if denominator_x <= 0 or denominator_y <= 0:
        return 0

    denominator = math.sqrt(denominator_x * denominator_y)

    if denominator == 0:
        return 0

    return (sum_xy - (sum_x * sum_y) / n) / denominator



def PearsonRecommend(nouveauCritique, critiques):
    """Propose une recommandation à Anne en utilisant le coefficient de corrélation de Pearson."""
    totals = {}
    sim_sums = {}

    # Parcourir tous les critiques
    for critique in critiques:
        if critique == nouveauCritique:
            continue

        # Calculer le coefficient de Pearson
        similarity = pearson(critiques[critique], critiques[nouveauCritique])

        # Ignorer les similarités nulles ou négatives
        if similarity <= 0:
            continue

        # Parcourir les films notés par ce critique
        for film in critiques[critique]:
            # Seuls les films non vus par Anne sont intéressants
            if film not in critiques[nouveauCritique] or critiques[nouveauCritique][film] == 0:
                # Ajouter le score pondéré au total
                if film not in totals:
                    totals[film] = 0
                totals[film] += critiques[critique][film] * similarity

                # Ajouter au dénominateur
                if film not in sim_sums:
                    sim_sums[film] = 0
                sim_sums[film] += similarity

    # Calculer la liste finale des recommandations
    rankings = [(total / sim_sums[film], film) for film, total in totals.items()]

    # Trier les recommandations par score décroissant
    rankings.sort(reverse=True)

    return rankings


# Test de la recommandation basée sur le coefficient de Pearson pour Anne
recommandations = PearsonRecommend('Anne', critiques)
print("Film recommandé à Anne avec le coefficient de Pearson :", recommandations[0])

# (4)

def cosinus_similarity(person1, person2):
    """Calcule la similarité cosinus en ignorant les valeurs 'None'."""
    communs = [film for film in person1 if film in person2 and person1[film] is not None and person2[film] is not None]

    if len(communs) == 0:
        return 0

    sum_xy = sum(person1[film] * person2[film] for film in communs)
    sum_x2 = sum(person1[film] ** 2 for film in communs)
    sum_y2 = sum(person2[film] ** 2 for film in communs)

    denominator = math.sqrt(sum_x2) * math.sqrt(sum_y2)
    if denominator == 0:
        return 0

    return sum_xy / denominator


def CosinusRecommend(nouveauCritique, critiques):
    """Propose une recommandation à Anne en utilisant la similarité cosinus."""
    totals = {}
    sim_sums = {}

    # Parcourir tous les critiques
    for critique in critiques:
        if critique == nouveauCritique:
            continue

        # Calculer la similarité cosinus
        similarity = cosinus_similarity(critiques[critique], critiques[nouveauCritique])

        # Ignorer les similarités nulles
        if similarity <= 0:
            continue

        # Parcourir les films notés par ce critique
        for film in critiques[critique]:
            # Seuls les films non vus par Anne sont intéressants
            if film not in critiques[nouveauCritique] or critiques[nouveauCritique][film] == 0:
                # Ajouter le score pondéré au total
                if film not in totals:
                    totals[film] = 0
                totals[film] += critiques[critique][film] * similarity

                # Ajouter au dénominateur
                if film not in sim_sums:
                    sim_sums[film] = 0
                sim_sums[film] += similarity

    # Calculer la liste finale des recommandations
    rankings = [(total / sim_sums[film], film) for film, total in totals.items()]

    # Trier les recommandations par score décroissant
    rankings.sort(reverse=True)

    return rankings


# Test de la recommandation basée sur la similarité cosinus pour Anne
recommandations = CosinusRecommend('Anne', critiques)
print("Film recommandé à Anne avec la similarité cosinus :", recommandations[0])

# (5)
# Générer des films et critiques
films = ['The Godfather', 'Pulp Fiction', 'Schindler\'s List', 'Casablanca', 'Citizen Kane',
         'The Shawshank Redemption', '12 Angry Men', 'Goodfellas', 'The Dark Knight',
         'Apocalypse Now', 'Forrest Gump', 'Inception', 'The Matrix', 'Star Wars', 'Jurassic Park',
         'Titanic', 'The Silence of the Lambs', 'Saving Private Ryan', 'The Lord of the Rings', 'Avatar']

critiques = ['Roger Ebert', 'Pauline Kael', 'A.O. Scott', 'Peter Travers', 'Leonard Maltin',
             'Gene Siskel', 'Richard Roeper', 'David Denby', 'Manohla Dargis', 'Anthony Lane',
             'Mark Kermode', 'James Berardinelli', 'Rex Reed', 'Christy Lemire', 'Owen Gleiberman']

def jaccard_similarity(person1, person2):
    """Calcule la similarité de Jaccard en considérant les films notés (ignorant les valeurs 'None')."""
    films_person1 = set(film for film in person1 if person1[film] is not None)
    films_person2 = set(film for film in person2 if person2[film] is not None)

    intersection = films_person1.intersection(films_person2)
    union = films_person1.union(films_person2)

    if len(union) == 0:
        return 0

    return len(intersection) / len(union)


def spearman_rank_correlation(person1, person2):
    """Calcule la corrélation de rang de Spearman en ignorant les valeurs 'None'."""
    communs = [film for film in person1 if film in person2 and person1[film] is not None and person2[film] is not None]

    if len(communs) < 2:
        return 0

    # Fonction pour calculer les rangs
    def rank_values(values):
        sorted_values = sorted(set(values))
        rank_dict = {value: rank for rank, value in enumerate(sorted_values, 1)}
        return [rank_dict[value] for value in values]

    # Calcul des rangs pour chaque personne
    rank1 = rank_values([person1[film] for film in communs])
    rank2 = rank_values([person2[film] for film in communs])

    # Calcul de la corrélation de Spearman
    n = len(communs)
    sum_d_squared = sum((r1 - r2) ** 2 for r1, r2 in zip(rank1, rank2))

    correlation = 1 - (6 * sum_d_squared) / (n * (n ** 2 - 1))

    return correlation


# Générer un tableau d'évaluations aléatoires avec entre 30% et 60% de cases vides
def generer_matrice_critique(films, critiques, utilisateur_cible, min_vide=0.3, max_vide=0.6):
    mat = {}
    for critique in critiques:
        mat[critique] = {}
        if critique == utilisateur_cible:
            # For the target user, randomly select up to 10 films to rate
            films_to_rate = random.sample(films, min(10, len(films)))
            for film in films:
                if film in films_to_rate:
                    mat[critique][film] = round(random.uniform(1, 5), 1)  # Note entre 1 et 5
                else:
                    mat[critique][film] = None
        else:
            for film in films:
                if random.random() > random.uniform(min_vide, max_vide):
                    mat[critique][film] = round(random.uniform(1, 5), 1)  # Note entre 1 et 5
                else:
                    mat[critique][film] = None
    return mat


# Calculer le pourcentage de cases vides
def pourcentagecasesvides(matrice):
    total_cases = len(matrice) * len(next(iter(matrice.values())))  # Nombre total de cases
    cases_vides = sum(
        1 for critique in matrice for film in matrice[critique] if matrice[critique][film] is None)  # Cases vides
    return (cases_vides / total_cases) * 100


# Fonction pour recommander des films basés sur une méthode de similarité
def recommander_film(nouveauCritique, critiques, methode_similarite):
    totals = {}
    sim_sums = {}
    for critique in critiques:
        if critique == nouveauCritique:
            continue
        similarity = methode_similarite(critiques[critique], critiques[nouveauCritique])
        if similarity <= 0:
            continue
        for film in critiques[critique]:
            # Vérifier si la note n'est pas None
            if critiques[critique][film] is not None and critiques[nouveauCritique][film] is None:
                if film not in totals:
                    totals[film] = 0
                totals[film] += critiques[critique][film] * similarity
                if film not in sim_sums:
                    sim_sums[film] = 0
                sim_sums[film] += similarity
    rankings = [(total / sim_sums[film], film) for film, total in totals.items()]
    rankings.sort(reverse=True)
    return rankings


# Fonction pour vérifier si toutes les recommandations sont identiques
def recommendations_identiques(recommandations):
    first_recommendation = recommandations[0][0][1] if recommandations[0] else None
    for rec in recommandations:
        if not rec or rec[0][1] != first_recommendation:
            return False
    return True


# Utilisateur cible pour la recommandation, on le fixe toujours à la derniere critique de la matrice
utilisateur_cible = 'Owen Gleiberman'

# Boucle pour générer des recommandations jusqu'à ce qu'elles soient toutes identiques
recommandations_identiques_flag = False
iteration = 0
while not recommandations_identiques_flag:
    iteration += 1
    matrice_critique = generer_matrice_critique(films, critiques, utilisateur_cible)

    # Calcul du pourcentage de cases vides
    pourcentage_vides = pourcentagecasesvides(matrice_critique)

    # Recommandations avec différentes méthodes de similarité (Manhattan, Euclidienne, Pearson, Cosinus, Jaccard,Spearman)
    recommandation_pearson = recommander_film(utilisateur_cible, matrice_critique, pearson)
    recommandation_cosinus = recommander_film(utilisateur_cible, matrice_critique, cosinus_similarity)
    recommandation_BestRecommend = Bestrecommend(utilisateur_cible, matrice_critique)
    recommandation_OtherBestRecommend = OtherBestrecommend(utilisateur_cible, matrice_critique)
    recommandation_jaccard = recommander_film(utilisateur_cible,matrice_critique,jaccard_similarity)
    recommandation_spearman = recommander_film(utilisateur_cible,matrice_critique,spearman_rank_correlation)


    # Stocker toutes les recommandations dans une liste
    toutes_recommandations = [
        recommandation_pearson,
        recommandation_cosinus,
        recommandation_BestRecommend,
        recommandation_OtherBestRecommend,
        recommandation_jaccard,
        recommandation_spearman
    ]

    # Vérifier si toutes les recommandations sont identiques
    recommandations_identiques_flag = recommendations_identiques(toutes_recommandations)


#exporter le result dans un fichier excel
def export_data_to_excel(matrice_critique, recommandations, utilisateur_cible, name):
    wb = Workbook()

    ws_matrice = wb.active
    ws_matrice.title = "Matrice Critique"

    ws_matrice.cell(row=1, column=1, value="Critiques / Films")
    for col, film in enumerate(next(iter(matrice_critique.values())).keys(), start=2):
        ws_matrice.cell(row=1, column=col, value=film)

    # Write data
    for row, (critique, films) in enumerate(matrice_critique.items(), start=2):
        ws_matrice.cell(row=row, column=1, value=critique)
        for col, (film, note) in enumerate(films.items(), start=2):
            ws_matrice.cell(row=row, column=col, value=note)

    # Create sheet for recommendations
    ws_recommend = wb.create_sheet(title="Recommendations")

    # Write recommendations
    ws_recommend.cell(row=1, column=1, value="Mesure de similarité")
    ws_recommend.cell(row=1, column=2, value="Film recommendé")
    ws_recommend.cell(row=1, column=3, value="Score")

    row = 2
    for measure, recommendation in recommandations.items():
        ws_recommend.cell(row=row, column=1, value=measure)
        ws_recommend.cell(row=row, column=2, value=recommendation[0][1])
        ws_recommend.cell(row=row, column=3, value=recommendation[0][0])
        row += 1

    # Create sheet for target user's ratings
    ws_target = wb.create_sheet(title="Notes de l'utilisateur cible")

    # Write target user's ratings
    ws_target.cell(row=1, column=1, value="Film")
    ws_target.cell(row=1, column=2, value="Note")

    row = 2
    for film, rating in matrice_critique[utilisateur_cible].items():
        ws_target.cell(row=row, column=1, value=film)
        ws_target.cell(row=row, column=2, value=rating)
        row += 1

    # Apply some basic styling
    for sheet in wb.worksheets:
        for cell in sheet["1:1"]:
            cell.font = Font(bold=True)
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Save the workbook
    wb.save(name + ".xlsx")


# Affichage des résultats finaux
print(f"\nNombre total d'itérations : {iteration}")
print(f"Pourcentage de cases vides : {pourcentage_vides:.2f}%")
print("\nRecommandations identiques trouvées :")
print("Recommandation (Manhattan):", recommandation_BestRecommend[0])
print("Recommandation (Euclidienne):", recommandation_OtherBestRecommend[0])
print("Recommandation (Pearson):", recommandation_pearson[0])
print("Recommandation (Cosinus):", recommandation_cosinus[0])
print("Recommandation (Jaccard):",recommandation_jaccard[0])
print("Recommandation (Spearman):",recommandation_spearman[0])
print(matrice_critique[utilisateur_cible])

recommandations = {
    "Manhattan": recommandation_BestRecommend,
    "Euclidienne": recommandation_OtherBestRecommend,
    "Pearson": recommandation_pearson,
    "Cosinus": recommandation_cosinus,
    "Jaccard": recommandation_jaccard,
    "Spearman": recommandation_spearman
}

export_data_to_excel(matrice_critique, recommandations, utilisateur_cible, "recommandation-identiques")


def explication_simple(film_recommande):
    explication = (
        f"Nous avons comparé vos goûts en matière de films avec ceux de personnes ayant des goûts similaires. "
        f"En fonction des films que vous avez aimés et des préférences de ces autres personnes, nous vous recommandons "
        f"le film '{film_recommande}', car ceux qui ont aimé les mêmes films que vous ont apprécié ce film."
    )
    return explication


# Génération d'une explication simple pour l'un des films recommandés
film_recommande = recommandation_pearson[0][1]
explication = explication_simple(film_recommande)
print("\nExplication pour l'utilisateur :")
print(explication)


# (6)

# Fonction pour vérifier si toutes les recommandations sont différentes
def recommendations_differentes(recommandations):
    films_recommandes = [rec[0][1] for rec in recommandations if
                         rec]  # Prendre le premier film recommandé de chaque méthode
    return len(set(films_recommandes)) == len(films_recommandes)  # Vérifier si tous les films sont distincts


# Boucle pour générer des recommandations jusqu'à ce qu'elles soient toutes différentes
recommandations_differentes_flag = False
iteration = 0
while not recommandations_differentes_flag:
    iteration += 1
    matrice_critique = generer_matrice_critique(films, critiques, utilisateur_cible)

    # Calcul du pourcentage de cases vides
    pourcentage_vides = pourcentagecasesvides(matrice_critique)

    # Recommandations avec différentes méthodes de similarité (Manhattan, Euclidienne, Pearson, Cosinus, Jaccard)
    recommandation_pearson = recommander_film(utilisateur_cible, matrice_critique, pearson)
    recommandation_cosinus = recommander_film(utilisateur_cible, matrice_critique, cosinus_similarity)
    recommandation_BestRecommend = Bestrecommend(utilisateur_cible, matrice_critique)
    recommandation_OtherBestRecommend = OtherBestrecommend(utilisateur_cible, matrice_critique)
    recommandation_jaccard = recommander_film(utilisateur_cible,matrice_critique,jaccard_similarity)
    recommandation_spearman = recommander_film(utilisateur_cible,matrice_critique,spearman_rank_correlation)

    # Stocker toutes les recommandations dans une liste
    toutes_recommandations = [
        recommandation_pearson,
        recommandation_cosinus,
        recommandation_BestRecommend,
        recommandation_OtherBestRecommend,
        recommandation_jaccard,
        recommandation_spearman
    ]

    # Vérifier si toutes les recommandations sont différentes
    recommandations_differentes_flag = recommendations_differentes(toutes_recommandations)

# Affichage des résultats finaux
print(f"\nNombre total d'itérations : {iteration}")
print(f"Pourcentage de cases vides : {pourcentage_vides:.2f}%")
print("\nRecommandations différentes :")
print("Recommandation (Manhattan):", recommandation_BestRecommend[0])
print("Recommandation (Euclidienne):", recommandation_OtherBestRecommend[0])
print("Recommandation (Pearson):", recommandation_pearson[0])
print("Recommandation (Cosinus):", recommandation_cosinus[0])
print("Recommandation (Jaccard):",recommandation_jaccard[0])
print("Recommandation (Spearman):",recommandation_spearman[0])

recommandations = {
    "Manhattan": recommandation_BestRecommend,
    "Euclidienne": recommandation_OtherBestRecommend,
    "Pearson": recommandation_pearson,
    "Cosinus": recommandation_cosinus,
    "Jaccard": recommandation_jaccard,
    "Spearman": recommandation_spearman
}

# Génération d'une explication simple pour l'un des films recommandés
film_recommande = recommandation_pearson[0][1]
explication = explication_simple(film_recommande)
print("\nExplication pour l'utilisateur :")
print(explication)

export_data_to_excel(matrice_critique, recommandations, utilisateur_cible, "recommandation-differentes")
